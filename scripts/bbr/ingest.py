# -*- coding: utf-8 -*-
"""
Læser Resights' BBR-tabeller og samler dem til én liste over lejligheder.

HVORFOR: opkøbs-tragtens trin 3 spørger hvor mange lejligheder der er i vores
størrelse. Det kunne ikke besvares, fordi vi kun kendte hver ejendoms samlede
kvm-SPÆND — og et spænd på 61-111 siger intet om hvor mange af 626 lejligheder
der ligger under 80. Kvm pr. lejlighed fandtes kun for de 73 vi selv ejer.

Én BBR-tabel dækker hele ejendommen. 36 filer dækker derfor alle 3.259 enheder,
hvor 3.259 enkeltopslag ville være uoverkommeligt.

BRUG:
    python3 scripts/bbr/ingest.py ~/Downloads
    python3 scripts/bbr/ingest.py ~/Downloads --ud src/lib/data/bbr-enheder.json

Uden --ud skrives intet; der vises kun hvad filerne indeholder. Det er med
vilje: man skal kunne se på tallene, før de bliver til data.

Kræver openpyxl. På denne maskine har system-python 3.9 den ikke — brug
/opt/homebrew/bin/python3.12.
"""
import sys, json, glob, os, re

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('openpyxl mangler. Prøv: /opt/homebrew/bin/python3.12 ' + ' '.join(sys.argv))

# Målgruppens grænser. Ændres de, ændres kun optællingen i oversigten —
# rådata i JSON'en er grænseløs, så den ikke skal genindlæses.
KVM_FRA, KVM_TIL = 20, 80


def _kol(header, *navne):
    """Find en kolonne på det første navn der findes. BBR-eksporter varierer."""
    for n in navne:
        if n in header:
            return header.index(n)
    return None


def _tekst(v):
    s = str(v or '').strip()
    return '' if s in ('-', 'None') else s


def laes_ejendom(sti):
    wb = load_workbook(sti, data_only=True)
    if 'Enheder' not in wb.sheetnames:
        return None

    stam = {}
    if 'Stamdata' in wb.sheetnames:
        for r in wb['Stamdata'].iter_rows(values_only=True):
            if r and r[0]:
                stam[str(r[0]).strip()] = r[1]

    link = str(stam.get('Link') or '')
    # BFE står sidst i Resights-linket; ellers står det i filnavnet.
    m = re.search(r'/(\d+)/?$', link) or re.search(r'BBRTabeller_(\d+)', os.path.basename(sti))
    ejendom_bfe = int(m.group(1)) if m else None

    ws = wb['Enheder']
    h = [c.value for c in ws[1]]
    ix = {
        'bfe':    _kol(h, 'BFE-nummer'),
        'vej':    _kol(h, 'Vejnavn'),
        'husnr':  _kol(h, 'Husnr.', 'Husnr'),
        'etage':  _kol(h, 'Etage'),
        'doer':   _kol(h, 'Dør'),
        'postnr': _kol(h, 'Postnr'),
        'by':     _kol(h, 'By'),
        'anv':    _kol(h, 'Enhedens anvendelse'),
        'kvm':    _kol(h, 'Enhedens samlede areal'),
        'kvmbeb': _kol(h, 'Areal til beboelse'),
        'vaer':   _kol(h, 'Antal værelser'),
        'status': _kol(h, 'Status'),
    }
    if ix['vej'] is None:
        return None

    enheder = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[ix['vej']]:
            continue
        etage, doer = _tekst(r[ix['etage']]), _tekst(r[ix['doer']])
        adr = f"{_tekst(r[ix['vej']])} {_tekst(r[ix['husnr']])}".strip()
        if etage or doer:
            adr += ', ' + f'{etage} {doer}'.strip()
        anv = _tekst(r[ix['anv']])
        kvm = r[ix['kvm']] if ix['kvm'] is not None else None
        enheder.append({
            'ejendomBfe': ejendom_bfe,
            'enhedBfe': r[ix['bfe']] if ix['bfe'] is not None else None,
            'adresse': adr,
            'postnr': _tekst(r[ix['postnr']]) if ix['postnr'] is not None else '',
            'by': _tekst(r[ix['by']]) if ix['by'] is not None else '',
            'anvendelse': anv,
            # BBR skelner ikke selv bolig fra erhverv i ét felt. Anvendelses-
            # teksten gør: alt der ikke starter med "Bolig" er kontor, værksted,
            # garage, kælder eller lignende — og skal ikke tælle som lejlighed.
            'erBolig': anv.startswith('Bolig'),
            # BBR fører også lejligheder der endnu ikke findes: «Projekteret»
            # og «Under opførelse», med 0 kvm og tom anvendelse. De må hverken
            # tælle som boliger eller som garager — de er ikke bygget endnu.
            'erOpfoert': _tekst(r[ix['status']]).startswith('Opført') if ix['status'] is not None else True,
            'kvm': kvm if isinstance(kvm, (int, float)) else None,
            'kvmBeboelse': r[ix['kvmbeb']] if ix['kvmbeb'] is not None else None,
            'vaerelser': r[ix['vaer']] if ix['vaer'] is not None else None,
            'status': _tekst(r[ix['status']]) if ix['status'] is not None else '',
        })

    return {
        'bfe': ejendom_bfe,
        'fil': os.path.basename(sti),
        'link': link,
        'opfoert': stam.get('Opførelsesår'),
        'antalBeboelse': stam.get('Antal beboelsesenheder'),
        'antalErhverv': stam.get('Antal erhvervsenheder'),
        'enheder': enheder,
    }


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    mappe = os.path.expanduser(sys.argv[1])
    ud = None
    if '--ud' in sys.argv:
        ud = sys.argv[sys.argv.index('--ud') + 1]

    filer = sorted(glob.glob(os.path.join(mappe, 'BBRTabeller*.xlsx')))
    if not filer:
        sys.exit(f'ingen BBRTabeller*.xlsx i {mappe}')

    ejendomme, set_bfe = [], set()
    for f in filer:
        e = laes_ejendom(f)
        if not e:
            print(f'  SPRINGER OVER (intet Enheder-ark): {os.path.basename(f)}')
            continue
        # Resights lægger "(1)"-kopier i Downloads. Samme BFE to gange ville
        # tælle hver lejlighed dobbelt.
        if e['bfe'] in set_bfe:
            print(f"  DUBLET, springes over: {e['fil']} (BFE {e['bfe']})")
            continue
        set_bfe.add(e['bfe'])
        ejendomme.append(e)

    print(f"\n{'ejendom':40} {'BFE':>9} {'enh':>4} {'bolig':>6} {'andet':>6} {'ubygget':>7} "
          f"{'i 20-80':>8} {'i 30-80':>8}  spænd")
    print('-' * 112)
    sum_ = dict(enh=0, bolig=0, andet=0, ubygget=0, m20=0, m30=0)
    for e in ejendomme:
        b = [u for u in e['enheder'] if u['erBolig'] and u['erOpfoert'] and u['kvm']]
        a = [u for u in e['enheder'] if not u['erBolig'] and u['erOpfoert']]
        ub = [u for u in e['enheder'] if not u['erOpfoert']]
        m20 = sum(1 for u in b if KVM_FRA <= u['kvm'] <= KVM_TIL)
        m30 = sum(1 for u in b if 30 <= u['kvm'] <= KVM_TIL)
        sp = f"{min(u['kvm'] for u in b):g}-{max(u['kvm'] for u in b):g}" if b else '—'
        navn = (e['link'].split('/')[-2] if e['link'].count('/') > 3 else e['fil'])[:40]
        print(f"{navn:40} {str(e['bfe']):>9} {len(e['enheder']):4} {len(b):6} "
              f"{len(a):6} {len(ub):7} {m20:8} {m30:8}  {sp}")
        sum_['enh'] += len(e['enheder']); sum_['bolig'] += len(b)
        sum_['andet'] += len(a); sum_['ubygget'] += len(ub)
        sum_['m20'] += m20; sum_['m30'] += m30
    print('-' * 112)
    print(f"{'I ALT — ' + str(len(ejendomme)) + ' ejendomme':40} {'':>9} "
          f"{sum_['enh']:4} {sum_['bolig']:6} {sum_['andet']:6} {sum_['ubygget']:7} "
          f"{sum_['m20']:8} {sum_['m30']:8}")

    if ud:
        alle = [u for e in ejendomme for u in e['enheder']]
        os.makedirs(os.path.dirname(ud) or '.', exist_ok=True)
        with open(ud, 'w', encoding='utf-8') as fh:
            json.dump({
                'kilde': f'Resights BBR-tabeller, {len(ejendomme)} ejendomme',
                'ejendomme': [{k: v for k, v in e.items() if k != 'enheder'} for e in ejendomme],
                'enheder': alle,
            }, fh, ensure_ascii=False, indent=1)
        print(f'\nskrevet: {ud}  ({len(alle)} enheder)')
    else:
        print('\n(intet skrevet — tilføj --ud <sti> når tallene ser rigtige ud)')


if __name__ == '__main__':
    main()
